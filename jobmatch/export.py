from __future__ import annotations

from collections import Counter
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from .models import MatchResult, SourceStatus


RESULT_COLUMNS = [
    "Rank",
    "Match score",
    "Skill score",
    "Role-title score",
    "Text similarity",
    "Experience score",
    "Location score",
    "Job title",
    "Company",
    "Location",
    "Country",
    "Work mode",
    "Seniority",
    "Job type",
    "Required years",
    "Language requirements",
    "Transferable matches",
    "Matched skills",
    "Missing skills",
    "Why it matches",
    "Warnings",
    "Official job URL",
    "Apply URL",
    "Provider",
    "Published/updated",
    "Application status",
    "Application date",
    "Notes",
    "Interview date",
    "Follow-up date",
]

TRACKER_COLUMNS = [
    "Job title",
    "Company",
    "Match score",
    "Application status",
    "Application date",
    "Notes",
    "Interview date",
    "Follow-up date",
    "Apply URL",
]

STATUS_OPTIONS = ("Not reviewed", "Saved", "Applied", "Rejected")


def _record_for(
    result: MatchResult,
    application_records: dict[str, dict[str, str]] | None,
) -> dict[str, str]:
    if not application_records:
        return {}
    key = (
        result.job.apply_url
        or result.job.job_url
        or f"{result.job.company}|{result.job.title}|{result.job.location}"
    )
    return application_records.get(key, {})


def results_to_dataframe(
    results: list[MatchResult],
    application_records: dict[str, dict[str, str]] | None = None,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for rank, result in enumerate(results, start=1):
        record = _record_for(result, application_records)
        rows.append(
            {
                "Rank": rank,
                "Match score": result.match_score,
                "Skill score": result.skill_score,
                "Role-title score": result.role_score,
                "Text similarity": result.text_score,
                "Experience score": result.experience_score,
                "Location score": result.location_score,
                "Job title": result.job.title,
                "Company": result.job.company,
                "Location": result.job.location,
                "Country": result.detected_country,
                "Work mode": result.work_mode,
                "Seniority": result.seniority,
                "Job type": result.job_type,
                "Required years": result.required_years,
                "Language requirements": ", ".join(result.language_requirements),
                "Transferable matches": ", ".join(result.transferable_skills),
                "Matched skills": ", ".join(result.matched_skills),
                "Missing skills": ", ".join(result.missing_skills),
                "Why it matches": " | ".join(result.reasons),
                "Warnings": " | ".join(result.warnings),
                "Official job URL": result.job.job_url,
                "Apply URL": result.job.apply_url or result.job.job_url,
                "Provider": result.job.provider,
                "Published/updated": result.job.published_at,
                "Application status": record.get("status", "Not reviewed"),
                "Application date": record.get("application_date", ""),
                "Notes": record.get("notes", ""),
                "Interview date": record.get("interview_date", ""),
                "Follow-up date": record.get("follow_up_date", ""),
            }
        )
    return pd.DataFrame(rows, columns=RESULT_COLUMNS)


def _missing_skills_dataframe(results: list[MatchResult]) -> pd.DataFrame:
    counts = Counter(skill for result in results for skill in result.missing_skills)
    total = len(results)
    return pd.DataFrame(
        [
            {
                "Missing skill": skill,
                "Jobs requiring it": count,
                "Share of jobs": count / total if total else 0.0,
            }
            for skill, count in counts.most_common()
        ],
        columns=["Missing skill", "Jobs requiring it", "Share of jobs"],
    )


def _source_status_dataframe(statuses: list[SourceStatus]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Company": status.company,
                "Provider": status.provider,
                "Status": status.status,
                "Jobs found": status.jobs_found,
                "Message": status.message,
            }
            for status in statuses
        ],
        columns=["Company", "Provider", "Status", "Jobs found", "Message"],
    )


def _method_dataframe() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("Generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")),
            ("Skill weight", "40%"),
            ("Role-title weight", "20%"),
            ("Text-similarity weight", "15%"),
            ("Experience weight", "15%"),
            ("Location weight", "10%"),
            (
                "Transferable skills",
                "Transferable engineering skills receive 65% of the skill component "
                "when both transferable and domain evidence are present.",
            ),
            (
                "Unknown metadata",
                "Unknown work mode, seniority, or job type is retained and flagged; "
                "it is not silently rejected.",
            ),
            (
                "Coverage boundary",
                "Only configured, verified official career feeds are searched. The "
                "Source Status sheet shows the actual coverage of this run.",
            ),
            (
                "Important",
                "Scores support human review; they are not hiring probabilities.",
            ),
        ],
        columns=["Item", "Value"],
    )


def _style_table_sheet(sheet, link_headers: tuple[str, ...] = ()) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_by_name: dict[str, int] = {}
    for cell in sheet[1]:
        header_by_name[str(cell.value)] = cell.column
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    narrow_headers = {
        "Rank",
        "Match score",
        "Skill score",
        "Role-title score",
        "Text similarity",
        "Experience score",
        "Location score",
        "Required years",
        "Jobs found",
        "Share of jobs",
    }
    wide_headers = {
        "Job title",
        "Location",
        "Language requirements",
        "Transferable matches",
        "Matched skills",
        "Missing skills",
        "Why it matches",
        "Warnings",
        "Official job URL",
        "Apply URL",
        "Notes",
        "Message",
        "Value",
    }
    for header, column in header_by_name.items():
        letter = sheet.cell(row=1, column=column).column_letter
        if header in narrow_headers:
            sheet.column_dimensions[letter].width = 15
        elif header in wide_headers:
            sheet.column_dimensions[letter].width = 38
        else:
            sheet.column_dimensions[letter].width = 20

    for row in range(2, sheet.max_row + 1):
        for header in link_headers:
            column = header_by_name.get(header)
            if column:
                cell = sheet.cell(row=row, column=column)
                if cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"
        for column in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=column).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    score_column = header_by_name.get("Match score")
    if score_column and sheet.max_row >= 2:
        letter = sheet.cell(row=1, column=score_column).column_letter
        sheet.conditional_formatting.add(
            f"{letter}2:{letter}{sheet.max_row}",
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="F8696B",
                mid_type="num",
                mid_value=60,
                mid_color="FFEB84",
                end_type="num",
                end_value=100,
                end_color="63BE7B",
            ),
        )


def build_excel(
    results: list[MatchResult],
    *,
    all_results: list[MatchResult] | None = None,
    source_statuses: list[SourceStatus] | None = None,
    application_records: dict[str, dict[str, str]] | None = None,
) -> bytes:
    """Create an auditable six-sheet job-search workbook."""

    all_results = all_results if all_results is not None else results
    source_statuses = source_statuses or []
    top_matches = results_to_dataframe(results, application_records)
    all_jobs = results_to_dataframe(all_results, application_records)
    tracker = all_jobs[TRACKER_COLUMNS].copy()

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        top_matches.to_excel(writer, index=False, sheet_name="Top Matches")
        all_jobs.to_excel(writer, index=False, sheet_name="All Jobs")
        _missing_skills_dataframe(all_results).to_excel(
            writer,
            index=False,
            sheet_name="Missing Skills",
        )
        tracker.to_excel(writer, index=False, sheet_name="Application Tracker")
        _source_status_dataframe(source_statuses).to_excel(
            writer,
            index=False,
            sheet_name="Source Status",
        )
        _method_dataframe().to_excel(writer, index=False, sheet_name="Scoring Method")

        for name in writer.book.sheetnames:
            links = ("Official job URL", "Apply URL") if name in {
                "Top Matches",
                "All Jobs",
                "Application Tracker",
            } else ()
            _style_table_sheet(writer.book[name], link_headers=links)

        tracker_sheet = writer.book["Application Tracker"]
        status_column = TRACKER_COLUMNS.index("Application status") + 1
        status_letter = tracker_sheet.cell(row=1, column=status_column).column_letter
        validation = DataValidation(
            type="list",
            formula1='"' + ",".join(STATUS_OPTIONS) + '"',
            allow_blank=False,
        )
        tracker_sheet.add_data_validation(validation)
        if tracker_sheet.max_row >= 2:
            validation.add(f"{status_letter}2:{status_letter}{tracker_sheet.max_row}")

        missing_sheet = writer.book["Missing Skills"]
        for row in range(2, missing_sheet.max_row + 1):
            missing_sheet.cell(row=row, column=3).number_format = "0.0%"

    return output.getvalue()
