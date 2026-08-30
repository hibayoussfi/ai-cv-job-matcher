from io import BytesIO

from openpyxl import load_workbook

from jobmatch.export import build_excel
from jobmatch.models import Job, MatchResult, SourceStatus


def test_excel_contains_six_sheets_ranked_job_and_clickable_link():
    job = Job(
        company="Example Energy",
        title="Battery Engineer",
        location="Berlin, Germany",
        description="BMS and Simulink",
        job_url="https://example.test/job",
        apply_url="https://example.test/job/apply",
        provider="Test",
    )
    result = MatchResult(
        job=job,
        match_score=82.5,
        skill_score=90,
        role_score=80,
        text_score=60,
        experience_score=100,
        location_score=100,
        matched_skills=("BMS", "Simulink"),
        missing_skills=("CAN",),
        required_years=3,
        detected_country="Germany",
    )
    statuses = [SourceStatus("Example Energy", "Test", "OK", jobs_found=1)]
    workbook = load_workbook(BytesIO(build_excel([result], source_statuses=statuses)))
    assert workbook.sheetnames == [
        "Top Matches",
        "All Jobs",
        "Missing Skills",
        "Application Tracker",
        "Source Status",
        "Scoring Method",
    ]
    sheet = workbook["Top Matches"]
    assert sheet["B2"].value == 82.5
    assert sheet["V2"].hyperlink.target == "https://example.test/job"
    assert sheet["Y2"].value == "Official company feed"
    assert workbook["Source Status"]["D2"].value == 1
    assert workbook["Scoring Method"]["B11"].value.startswith("Scores support human review")


def test_excel_tracker_includes_session_status_and_notes():
    job = Job(
        company="Example",
        title="Controls Engineer",
        location="Munich, Germany",
        description="Simulink",
        job_url="https://example.test/job",
        apply_url="https://example.test/apply",
    )
    result = MatchResult(
        job=job,
        match_score=70,
        skill_score=80,
        role_score=60,
        text_score=50,
        experience_score=70,
        location_score=100,
    )
    records = {
        "https://example.test/apply": {
            "status": "Saved",
            "notes": "Review requirements",
        }
    }
    workbook = load_workbook(
        BytesIO(build_excel([result], application_records=records))
    )
    tracker = workbook["Application Tracker"]
    assert tracker["D2"].value == "Saved"
    assert tracker["F2"].value == "Review requirements"
    assert tracker["I2"].hyperlink.target == "https://example.test/apply"
